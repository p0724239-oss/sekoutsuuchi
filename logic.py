import calendar
import re
from io import BytesIO

import pandas as pd
from openpyxl.styles import PatternFill


def _strip_eq(val):
    if pd.isna(val):
        return val
    m = re.match(r'^="?(.+?)"?$', str(val).strip())
    return m.group(1) if m else str(val).strip()


def _norm_num(v):
    try:
        return str(int(float(str(v))))
    except Exception:
        return str(v).strip()


def load_kwzweb(file) -> pd.DataFrame:
    df = pd.read_csv(file, encoding="cp932", skiprows=2, dtype=str)
    for col in df.columns:
        df[col] = df[col].apply(_strip_eq)

    df["_ord"] = df["オーダ番号"].apply(_norm_num)
    df["_cls"] = df["工事分類"].str.strip()
    df["_num"] = df["施行通知書番号"].apply(_norm_num)
    df["_枝番号"] = pd.to_numeric(df["施行通知書枝番号"], errors="coerce")
    df["_最大行番号"] = pd.to_numeric(df["最大行番号"], errors="coerce")
    df["_最終更新日"] = pd.to_datetime(df["最終更新日"], errors="coerce")

    return df


def load_jisseki(file) -> pd.DataFrame:
    df = pd.read_excel(file, sheet_name="作業実績ファイル（労務費）", dtype=str)
    df = df[df["主工種フラグ"] == "1"].copy()

    df["_ord"] = df["オーダ番号"].apply(_norm_num)
    df["_cls"] = df["工事分類名称"].str.strip()
    df["_num"] = df["施行通知書番号"].apply(_norm_num)
    df["_行番号"] = pd.to_numeric(df["施行通知書行番号"], errors="coerce")
    df["_施行日"] = pd.to_datetime(df["施行日"], errors="coerce")

    return df


def _calc_deadline(施行日, deadline_day: int, months_before: int):
    if pd.isna(施行日):
        return pd.NaT
    try:
        target = 施行日 - pd.DateOffset(months=months_before)
        last_day = calendar.monthrange(target.year, target.month)[1]
        day = min(deadline_day, last_day)
        return pd.Timestamp(year=target.year, month=target.month, day=day)
    except Exception:
        return pd.NaT


def analyze(
    kwzweb_df: pd.DataFrame,
    jisseki_df: pd.DataFrame,
    deadline_day: int = 15,
    months_before: int = 1,
) -> tuple[pd.DataFrame, int]:
    # kwzweb を結合キーでグループ化（枝番号昇順でソート済み）
    kw_grouped = {
        key: grp.sort_values("_枝番号").reset_index(drop=True)
        for key, grp in kwzweb_df.groupby(["_ord", "_cls", "_num"])
    }

    records = []
    skipped = 0

    for _, row in jisseki_df.iterrows():
        key = (row["_ord"], row["_cls"], row["_num"])
        branches = kw_grouped.get(key)

        if branches is None:
            skipped += 1
            continue

        行番号 = row["_行番号"]
        施行日 = row["_施行日"]
        期日 = _calc_deadline(施行日, deadline_day, months_before)

        # 行番号が最初に含まれる枝番（最大行番号 >= 行番号 の最小枝番）
        valid = branches[branches["_最大行番号"] >= 行番号]

        if valid.empty:
            判定 = "NG（枝番不明）"
            branch_row = pd.Series(dtype=object)
        else:
            branch_row = valid.iloc[0]
            ステータス = branch_row["施行通知ステータス"]
            発行日 = branch_row["_最終更新日"]

            if ステータス != "承認済":
                判定 = f"NG（{ステータス}）"
            elif pd.isna(発行日) or pd.isna(期日):
                判定 = "NG（日付不明）"
            elif 発行日.normalize() <= 期日:
                判定 = "OK"
            else:
                判定 = "NG（期日超過）"

        records.append(
            {
                "箇所名": row.get("箇所名", ""),
                "オーダ番号": row["オーダ番号"],
                "工事番号": row.get("工事番号", ""),
                "工事件名": row.get("工事件名", ""),
                "工事分類": row["工事分類名称"],
                "施行通知書番号": row["施行通知書番号"],
                "施行通知書行番号": row["施行通知書行番号"],
                "工種名称": row.get("工種名称", ""),
                "請負者名": row.get("請負者名", ""),
                "施行日": 施行日,
                "期日": 期日,
                "発行日": branch_row.get("_最終更新日", pd.NaT) if not branch_row.empty else pd.NaT,
                "判定": 判定,
                "枝番号": branch_row.get("施行通知書枝番号", "") if not branch_row.empty else "",
                "通知ステータス": branch_row.get("施行通知ステータス", "") if not branch_row.empty else "",
            }
        )

    return pd.DataFrame(records), skipped


def to_excel(result_df: pd.DataFrame) -> bytes:
    total = len(result_df)
    ok_count = (result_df["判定"] == "OK").sum()
    ng_count = total - ok_count

    def fmt_dates(df: pd.DataFrame) -> pd.DataFrame:
        d = df.copy()
        for col in ["施行日", "期日"]:
            d[col] = pd.to_datetime(d[col]).dt.strftime("%Y/%m/%d")
        d["発行日"] = pd.to_datetime(d["発行日"]).dt.strftime("%Y/%m/%d %H:%M")
        return d

    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        # 集計シート
        summary = pd.DataFrame(
            {
                "項目": ["集計対象件数", "OK件数", "NG件数", "OK率"],
                "値": [
                    total,
                    ok_count,
                    ng_count,
                    f"{ok_count / total * 100:.1f}%" if total > 0 else "0%",
                ],
            }
        )
        summary.to_excel(writer, sheet_name="集計", index=False)

        # 全件明細シート（色付き）
        detail = fmt_dates(result_df)
        detail.to_excel(writer, sheet_name="全件明細", index=False)

        ws = writer.sheets["全件明細"]
        green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        n_cols = len(detail.columns)
        for i, judge in enumerate(detail["判定"], start=2):
            fill = green if judge == "OK" else red
            for c in range(1, n_cols + 1):
                ws.cell(row=i, column=c).fill = fill

        # NG一覧シート
        ng_df = fmt_dates(result_df[result_df["判定"] != "OK"])
        ng_df.to_excel(writer, sheet_name="NG一覧", index=False)

    return output.getvalue()
